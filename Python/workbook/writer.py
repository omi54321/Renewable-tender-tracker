from __future__ import annotations

import shutil
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook.properties import CalcProperties

MASTER_HEADERS = [
    "Tender ID","Category","Agency","State","Tender Number","Technology","Capacity MW",
    "Storage Required","Storage Hours","Dispatch Condition","Minimum PLF %","EMD",
    "BG / PBG Conditions","Bid Submission Date","Technical Opening","Financial Opening",
    "Status","Last Updated","Corrigendum Count","Tender URL","Document URL","Remarks",
    "Days to Deadline","Deadline Alert","Duplicate Key","Duplicate Flag","Change Fingerprint",
    "Verified On","Source Quality","Data Completeness","Evidence Score","Opportunity Class",
    "Tender Notification Date","Corrigendum Date(s)","Latest Corrigendum Date","Corrigendum Details",
    "Procurement Model","Supply Window / Dispatch Selection","Daily Supply Obligation",
    "Monthly Minimum / Reconciliation","Annual PLF / CUF / Energy Requirement",
    "Allowed Shortfall / Availability Tolerance","Under-Supply Penalty",
    "External / Exchange Energy Allowed","Maximum External Energy",
    "Charging / External Energy Conditions","Third-Party / Merchant Use",
    "Detailed Source Document","Technical Review Status",
    "First Seen By Automation","Last Seen By Automation","Source Record Hash",
]

def backup(workbook_path: Path, backup_dir: Path) -> Path:
    backup_dir.mkdir(parents=True, exist_ok=True)
    target = backup_dir / f"{workbook_path.stem}_{datetime.now():%Y%m%d_%H%M%S}{workbook_path.suffix}"
    shutil.copy2(workbook_path, target)
    return target

def read_master(workbook_path: Path) -> list[dict[str, Any]]:
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")
    workbook = load_workbook(workbook_path, data_only=False)
    sheet = workbook["Master Tenders"]
    headers = {sheet.cell(1, col).value: col for col in range(1, sheet.max_column + 1)}
    records = []
    for row in range(2, sheet.max_row + 1):
        if not sheet.cell(row, headers.get("Tender ID", 1)).value:
            continue
        records.append({h: sheet.cell(row, headers[h]).value for h in MASTER_HEADERS if h in headers})
    return records

def _copy_style(source, target) -> None:
    if source.has_style: target._style = copy(source._style)
    if source.number_format: target.number_format = source.number_format
    target.alignment = copy(source.alignment)
    target.protection = copy(source.protection)

def write_master(workbook_path: Path, active_records: list[dict[str, Any]], exclusions: list[dict[str, Any]], run_result: str) -> None:
    backup(workbook_path, Path("backups"))
    workbook = load_workbook(workbook_path)
    master = workbook["Master Tenders"]
    header_map = {master.cell(1, col).value: col for col in range(1, master.max_column + 1)}

    # Preserve formatting while clearing active values.
    for row in range(2, max(master.max_row, 501) + 1):
        for header in MASTER_HEADERS:
            col = header_map.get(header)
            if col: master.cell(row, col).value = None

    now = datetime.now()
    for row_index, record in enumerate(active_records, start=2):
        record.setdefault("Last Updated", now)
        record.setdefault("First Seen By Automation", now)
        record["Last Seen By Automation"] = now
        for header in MASTER_HEADERS:
            col = header_map.get(header)
            if col:
                if row_index > 2: _copy_style(master.cell(2, col), master.cell(row_index, col))
                master.cell(row_index, col).value = record.get(header)

        # Formula columns
        if "Days to Deadline" in header_map:
            master.cell(row_index, header_map["Days to Deadline"]).value = f'=IF(N{row_index}="","",N{row_index}-TODAY())'
        if "Deadline Alert" in header_map:
            master.cell(row_index, header_map["Deadline Alert"]).value = f'=IF(W{row_index}="","",IF(W{row_index}<0,"Expired",IF(W{row_index}<=3,"Urgent",IF(W{row_index}<=7,"Due in 7 Days","Normal"))))'
        if "Duplicate Key" in header_map:
            master.cell(row_index, header_map["Duplicate Key"]).value = f'=UPPER(TRIM(C{row_index}&"|"&E{row_index}))'
        if "Duplicate Flag" in header_map:
            master.cell(row_index, header_map["Duplicate Flag"]).value = f'=IF(Y{row_index}="","",IF(COUNTIF($Y$2:$Y$501,Y{row_index})>1,"Duplicate","Unique"))'

    if "Exclusion Log" in workbook.sheetnames:
        exclusion_sheet = workbook["Exclusion Log"]
        for record in exclusions:
            exclusion_sheet.append([
                record.get("Tender ID", ""), record.get("Agency", ""), record.get("Tender Number", ""),
                record.get("Technology", ""), record.get("Capacity MW", ""), record.get("Status", ""),
                record.get("Exclusion Reason", ""), record.get("Tender URL", ""),
            ])

    if "Automation Status" in workbook.sheetnames:
        status = workbook["Automation Status"]
        status["C13"] = now
        status["C14"] = run_result
        status.append([None,None,None,None,None,f"RUN-{now:%Y%m%d%H%M%S}",now,now,len(active_records),run_result])

    if workbook.calculation is None:
        workbook.calculation = CalcProperties()
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcOnSave = True
    workbook.save(workbook_path)
